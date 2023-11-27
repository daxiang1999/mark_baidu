import pandas as pd
import numpy as np
import win32com.client as win32
import shutil
import os
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import load_workbook, Workbook
from copy import copy
import xlwings as xw
import re


def fill_down(df, columns, method='ffill') -> pd.DataFrame:
    """
    对某一列或几列的空字符串或者是nan进行填充,columns是一个列表,包括需要进行填充的列名
    df: 需要进行填充的df
    columns: 需要进行填充的列名
    method: 填充的方法，ffill是用该列中上一个非缺失值进行填充。bfill是用该列中下一个非缺失值进行填充
    """
    def change(x):
        if not pd.isna(x):  # 如果不是空值
            if str(x).replace(' ', '') == '':
                return pd.NA
            else:
                return x
        else:
            return x

    if isinstance(columns, str):
        columns = [columns]

    for col in columns:
        df[col] = df[col].apply(change).fillna(method=method)
    return df


def remove_not_last(df, gsmc='公司名称', jznd='记账年度', kmdm='科目代码') -> pd.DataFrame:
    """删除非末级科目"""
    need_remove_index = []
    for step, i in enumerate(df.index):
        length = len(df.loc[i, kmdm])
        if i != max(df.index):
            if ((df.loc[i, gsmc] + df.loc[i, jznd] + df.loc[i, kmdm] ==
                    df.loc[df.index[step + 1], gsmc] + df.loc[df.index[step + 1], jznd] + df.loc[df.index[step + 1], kmdm][0: length]) and
                    (df.loc[i, gsmc] + df.loc[i, jznd] + df.loc[i, kmdm] !=
                     df.loc[df.index[step + 1], gsmc] + df.loc[df.index[step + 1], jznd] + df.loc[df.index[step + 1], kmdm])):
                need_remove_index.append(i)
    df.drop(index=need_remove_index, inplace=True)
    return df


def get_kmdm_dict(df, kmdm='科目代码', kmmc='科目名称', sep='-') -> dict:
    """对科目名称进行拆分，并重新组合成包含多级科目名称形式"""
    df = df.loc[:, [kmdm, kmmc]]
    df.loc[:, kmdm] = df.loc[:, kmdm].astype(str)
    df.loc[:, kmmc] = df.loc[:, kmmc].apply(lambda x: str(x).split(sep)[-1].strip())
    kmdm_ori_dict = df.drop_duplicates().set_index(kmdm)[kmmc].to_dict()
    kmdm_new_dict = {}
    for key, value in kmdm_ori_dict.items():
        new_value = ''
        for ele in kmdm_ori_dict.keys():
            if key.startswith(ele) and key != ele:
                new_value += kmdm_ori_dict.get(ele) + '-'
        new_value = new_value + value
        kmdm_new_dict.update({key: new_value})
    return kmdm_new_dict


def read(io, **kwargs) -> pd.DataFrame:
    """通用的读表函数，可以根据文件后缀名分别读取excel或者csv"""
    if io.endswith('xlsx'):
        try:
            data = pd.read_excel(io, **kwargs)
        except UnicodeDecodeError:
            kwargs.update({'encoding': 'gbk'})
            data = pd.read_excel(io, **kwargs)
    elif io.endswith('csv'):
        if 'low_memory' not in kwargs.keys():
            kwargs.update({'low_memory': False})
        try:
            data = pd.read_csv(io, **kwargs)
        except UnicodeDecodeError:
            kwargs.update({'encoding': 'gbk'})
            data = pd.read_csv(io, **kwargs)
    else:
        raise ValueError(f'【{io}】不是excel或者csv文件，请检查文件格式后重试。')
    return data


def translate_to_xlsx(input_dir, output_dir):
    """将一个文件下的文件全部转化成xlsx格式，并另存到另一个文件夹下"""
    if not os.path.exists(input_dir):
        raise FileNotFoundError('输入文件夹不存在，请检查后重试。')
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)  # 可以创建多层目录
    excel = win32.DispatchEx('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    for file in os.listdir(input_dir):
        if file.endswith(".xls") or file.endswith('.XLS'):
            io = os.path.join(input_dir, file)
            wb = excel.Workbooks.Open(io)
            file = file.replace('.xls', '').replace('.XLS', '')
            full_path = os.path.join(output_dir, file)
            if '.' in full_path:
                full_path += '.xlsx'
            wb.SaveAs(full_path, FileFormat=51)
            wb.Close()
        elif file.endswith('.xlsx') or file.endswith('.XLSX'):
            source_file = os.path.join(input_dir, file)
            target_file = os.path.join(output_dir, file)
            shutil.copy(source_file, target_file)
        else:
            continue


def handle_kmb(xsz_df, yeb_df, wlb_df, sep='-') -> pd.DataFrame:
    """生成科目标识表，ETL字典表和会计科目初步标识表的文件路径需要换成自己的"""
    num_dict = ['一', '二', '三', '四', '五', '六', '七', '八', '九', '十']
    xsz_df = xsz_df[['科目代码', '科目名称']]
    yeb_df = yeb_df[['科目代码', '科目名称']]
    wlb_df = wlb_df[['科目代码', '科目名称']]
    df = pd.concat([xsz_df, yeb_df, wlb_df], axis=0)
    df['科目代码'] = df['科目代码'].astype(str)
    df.drop_duplicates(inplace=True)

    for i, num in enumerate(num_dict):
        # 增加科目层级名称,科目名称之间的分隔符为sep
        df[f'{num}级科目层级名称'] = df['科目名称'].apply(lambda x: x.split(sep)[i] if len(x.split(sep)) > i else '')
        # 增加全级科目层级名称
        df[f'{num}级科目全级名称'] = df['科目名称'].apply(lambda x: sep.join(x.split(sep)[0:i + 1]) if len(x.split(sep)) > i else '')

    # 科目层级
    df['科目层级'] = df['科目名称'].map(lambda x: len(x.split(sep)))

    # 但是科目代码之间的分隔符，没有办法确定的，需要先判断，一般常见的科目代码分隔符有三种 . - _
    kmdm_sep = None
    for ele in '.-_':
        if df['科目代码'].str.contains(ele, regex=False).any():
            kmdm_sep = ele

    # 根据科目代码的分隔符来划分【科目层级代码】
    if kmdm_sep:
        for i, num in enumerate(num_dict):
            # 增加层级科目代码
            df[f'{num}级科目层级代码'] = df['科目代码'].astype(str).map(lambda x: x.split(kmdm_sep)[i] if len(x.split(kmdm_sep)) > i else '')
            # 增加全级科目代码
            df[f'{num}级科目全级代码'] = df['科目代码'].astype(str).map(lambda x: kmdm_sep.join(x.split(kmdm_sep)[0:i + 1]) if len(x.split(kmdm_sep)) > i else '')

    else:
        # 如果没有分隔符，默认科目代码层级划分是4-2-2-2-2的形式 // 判断一下公司代码的长度有些是4-3-3-3的形式
        # kmdm_level_len_list = df['科目代码'].apply(lambda ).unique().tolist()
        for i, num in enumerate(num_dict):
            # 增加全级科目代码
            df[f'{num}级科目全级代码'] = df['科目代码'].astype(str).apply(lambda x: str(x)[0:2 * i + 4] if len(str(x)) > 2 * i + 3 else '')
            # 增加层级科目代码
            if i == 0:
                df['一级科目层级代码'] = df['科目代码'].astype(str).str[0:4]
            else:
                df[f'{num}级科目层级代码'] = df['科目代码'].apply(lambda x: str(x)[2 * i + 2:2 * i + 4] if len(str(x)) > 2 * i + 3 else '')

    # 添加一些其他的列名以及缺失的列名
    df['科目代码-复制'] = df['科目代码']
    sawp_da = r"E:\TJ_DA\ETL清洗\ETL数据处理模板\SAWP_DA_业财数据ETL字典表_230601.xlsx"
    da_df = pd.read_excel(sawp_da, sheet_name='科目标识表')
    for col in da_df.columns:
        if col not in df.columns:
            df[col] = ''

    # 按照一级层级匹配【科目报表标识】【科目借贷标识】【科目类型标识】
    template_path = r"E:\TJ_DA\ETL清洗\ETL数据处理模板\会计科目初步标识表.xlsx"
    standard_df = pd.read_excel(template_path)
    kmjdbs_dict = standard_df.set_index('kmmc_part_level1')['kmjdbs'].to_dict()
    kmlxbs_dict = standard_df.set_index('kmmc_part_level1')['kmlxbs'].to_dict()
    kmbbbs_dict = standard_df.set_index('kmmc_part_level1')['kmbbbs'].to_dict()

    df['科目借贷标识'] = df['一级科目层级名称'].map(kmjdbs_dict)
    df['科目类型标识'] = df['一级科目层级名称'].map(kmlxbs_dict)
    df['科目报表标识'] = df['一级科目层级名称'].map(kmbbbs_dict)

    # 检测科目报表标识、科目借贷标识的完整度
    if df['科目借贷标识'].isna().sum() != 0:
        lack_list = df.loc[df['科目借贷标识'].isna(), '一级科目全级名称'].unique().tolist()
        print('在会计科目初步标识表中没有匹配到一下科目名称：请手动补充！')
        print(lack_list)
    return df


def add_pzhh(df, gsmc='公司名称', jznd='记账年度', jzyd='记账月度', pzzh='凭证字号', pzbh='凭证编号', pzhh='凭证行号') -> pd.DataFrame:
    """给序时账增加凭证行号"""
    try:
        df = df.sort_values(by=[gsmc, jznd, jzyd, pzzh, pzbh])
    except KeyError:
        print('【公司名称、记账年度、记账月度、凭证字号、凭证编号】这几列是否都存在，请检查后重试。')
        return
    df[pzhh] = ''
    for step, i in enumerate(df.index):
        if i == df.index[0]:
            df.loc[i, pzhh] = 1
        else:
            if ((df.loc[i, gsmc] == df.loc[df.index[step - 1], gsmc]) and (df.loc[i, jznd] == df.loc[df.index[step - 1], jznd]) and
                    (df.loc[i, jzyd] == df.loc[df.index[step - 1], jzyd]) and (df.loc[i, pzzh] == df.loc[df.index[step - 1], pzzh])
                    and (df.loc[i, pzbh] == df.loc[df.index[step - 1], pzbh])):
                # 公司名称、记账年度、记账月度、凭证字号、凭证编号相同的情况下，凭证行号+1
                df.loc[i, pzhh] = df.loc[df.index[step - 1], pzhh] + 1
            else:
                df.loc[i, pzhh] = 1
    return df


def insert_rows(df: pd.DataFrame, index: int = None, count: int = 1) -> pd.DataFrame:
    """在df的index位置插入count行,默认在最后一行插入一行"""
    if index is None:
        index = df.index[-1] + 1
    df = df.reset_index(drop=True)
    dfs = np.split(df, [index])   # 前index行为一个DataFrame,后面的为一个DataFrame
    new_df = pd.DataFrame([[np.nan]*len(df.columns) for _ in range(count)], columns=df.columns)
    dfs.insert(1, new_df)
    df = pd.concat(dfs, ignore_index=True)
    return df


def reconstruction_columns(df, seq='_') -> pd.DataFrame:
    """重构列名,这里的df的列名是多重行标题"""
    new_columns = []
    for cols in df.columns:
        temp = [x for x in cols if 'Unnamed' not in x]
        new_columns.append(seq.join(temp))

    df.columns = new_columns
    return df


def sht_copy(src_wb, tar_wb, src_ws, tar_ws=None, index=-1):
    """
    复制一个工作表到另一个工作簿
    src_wb: 待复制的工作表所属的工作簿,可以传入workbook对象或者是文件路径
    tar_wb: 目标工作簿,可以传入workbook对象或者是文件路径
    src_ws: 待复制的工作表表名
    tar_ws: 目标工作表名称，用于复制过后对工作表重命名
    """
    # 处理输入的src_wb，tar_wb
    # 如果是字符串则通过openpyxl打开工作簿
    if isinstance(src_wb, str):
        src_wb = load_workbook(src_wb, data_only=True)
    if isinstance(tar_wb, str):
        tar_wb = load_workbook(tar_wb)
    # 处理输入的tar_ws 如果没有传入目标worksheet名称，则使用源worksheet名称
    if tar_ws is None:
        tar_ws = src_ws
    # 新建目标worksheet和打开源worksheet
    new_sheet = tar_wb.create_sheet(title=tar_ws, index=index)  # 在文件最后面新建一个目标sheet
    ori_sheet = src_wb[src_ws]

    # 第1步：复制非合并单元格
    for row in range(1, ori_sheet.max_row + 1):
        for col in range(1, ori_sheet.max_column + 1):
            # 复制列宽：
            new_sheet.column_dimensions[get_column_letter(col)].width = \
                ori_sheet.column_dimensions[get_column_letter(col)].width
            # 复制值：
            new_sheet.cell(row=row, column=col).value = ori_sheet.cell(row=row, column=col).value
            # 如有格式则复制格式：
            if ori_sheet.cell(row=row, column=col).has_style:
                new_sheet.cell(row=row, column=col).font = copy(ori_sheet.cell(row=row, column=col).font)
                new_sheet.cell(row=row, column=col).border = copy(ori_sheet.cell(row=row, column=col).border)
                new_sheet.cell(row=row, column=col).fill = copy(ori_sheet.cell(row=row, column=col).fill)
                new_sheet.cell(row=row, column=col).alignment = copy(ori_sheet.cell(row=row, column=col).alignment)
                new_sheet.cell(row=row, column=col).number_format = copy(
                    ori_sheet.cell(row=row, column=col).number_format)

    # 第2步：重新合并单元格
    for i in ori_sheet.merged_cells:
        new_sheet.merge_cells(i.coord)

    return tar_wb


def find_one_kind_file(dir, feature, mode='startswith'):
    """查找dir文件夹下所有以feature开头/结尾的文件, mode可以是startswith或者是endswith"""
    if mode == 'startswith':
        syn = 'file.startswith(feature)'
    elif mode == 'endswith':
        syn = 'file.endswith(feature)'
    else:
        raise ValueError('mode参数只能是startswith或者是endswith')
    file_list = []
    for file in os.listdir(dir):
        file_path = os.path.join(dir, file)
        if os.path.isfile(file_path):
            if eval(syn):
                file_list.append(file_path)
        else:
            file_list.extend(find_one_kind_file(file_path, feature, mode))
    return file_list


def insert_sheet(ori_path, ori_sheetname, tar_path, index=-1):
    """将一个工作表插入到另一个工作簿中,暂时没有什么用，因为每次都要打开关闭kill掉excel，速度太慢了"""
    app = xw.App(visible=False)
    app.display_alerts = False  # 保存提示
    app.screen_updating = False  # 屏幕刷新
    # 加载需要插入的工作表
    wb = app.books.open(ori_path)
    ws = wb.sheets[ori_sheetname]
    # 加载目标工作簿
    tar_wb = app.books.open(tar_path)
    # 将工作表插入到目标工作簿
    ws.api.Copy(Before=tar_wb.sheets[index].api)
    wb.close()
    tar_wb.save()
    tar_wb.close()
    app.quit()
    app.kill()


def write_df(ws, data, range_string=None, row=None, col=None):
    """
    openpyxl打开的工作表写入df数据,包括表头和数据
    """
    if range_string is not None:
        range_s = re.findall(re.compile('[a-zA-Z]+'), range_string)[0]
        range_id = re.findall(re.compile(r'\d+'), range_string)[0]
        start_row = int(range_id)
        start_col = column_index_from_string(range_s)
    elif row is not None and col is not None:
        start_row = row
        start_col = col
    else:
        raise ValueError("range_string and (row, col) cannot both be None")
    end_row = start_row + data.shape[0] + 1
    end_col = start_col + data.shape[1]
    data.fillna('', inplace=True)
    # 把dataframe中的表头写入到工作表中
    for c in range(start_col, end_col):
        ws.cell(start_row, c).value = data.columns[c - start_col]
    # 把dataframe中的数据写入到工作表中
    for r in range(start_row+1, end_row):
        for c in range(start_col, end_col):
            ws.cell(r, c).value = data.iloc[r - start_row - 1, c - start_col]


def set_number_format(ws, start_row, start_col, end_row=None, end_col=None, number_format=None):

    if end_row is None:
        end_row = ws.max_row
    if end_col is None:
        end_col = ws.max_column
    if number_format is None:
        number_format = '#,##0.00_);(#,##0.00)'
    for i in range(start_row, end_row+1):
        for j in range(start_col, end_col+1):
            ws.cell(i, j).number_format = number_format


def group_sum(df, col, cols=None):
    """对原DataFrame根据col分组，然后对每组的数字列求和，最后将每组的和与原DataFrame合并"""
    # 找出所有DataFrame的数字列
    if cols is None:
        cols = df.select_dtypes(include='number').columns.tolist()
    sum_df = df.groupby(col).agg(dict.fromkeys(cols, 'sum'))
    total_list = []
    for name, group in df.groupby(col):
        group = pd.concat([group, sum_df.loc[name]], axis=0)
        total_list.append(group)

    return pd.concat(total_list)


if __name__ == '__main__':
    # 方法一调用xlwings写插入表，会出bug，插入的位置不对
    # app = xw.App(visible=False)
    # app.display_alerts = False  # 保存提示
    # app.screen_updating = False  # 屏幕刷新
    # path_dir = r'E:\projects\九总七部_230831_晋亿公司年报审计\05_输出的_底稿文件\九总七部_230831_晋亿公司年报审计_SAWP_230912123024\不合并'
    # file_list = find_one_kind_file(path_dir, '5', mode='startswith')
    # # 加载需要插入的工作表
    # inserted_wb = app.books.open(r"C:\Users\xuxx0\Desktop\text.xlsx")
    # inserted_ws = inserted_wb.sheets['Sheet1']
    # for file in file_list:
    #     tar_wb = app.books.open(file)
    #     # 计算tar_wb中的工作表数量
    #     tar_ws_num = len(tar_wb.sheets)
    #     # 将inserted_ws插入到tar_wb的最后一个工作表之后
    #     inserted_ws.api.Copy(Before=tar_wb.sheets[tar_ws_num-1].api)
    #     tar_wb.save()
    #     tar_wb.close()
    #     print(f'{file}已经插入了一张新表')
    # inserted_wb.close()
    # app.quit()
    # app.kill()

    # 方法二： 使用openpyxl来插入表
    path_dir = r'E:\projects\九总七部_230831_晋亿公司年报审计\05_输出的_底稿文件\九总七部_230831_晋亿公司年报审计_SAWP_230912123024\不合并'
    file_list = find_one_kind_file(path_dir, '6', mode='startswith')
    # 需要插入的工作表路径和表名
    src_wb = load_workbook(r"C:\Users\xuxx0\Desktop\result.xlsx")
    src_ws = '固定资产明细表'
    for file in file_list:
        wb = sht_copy(src_wb, file, src_ws)
        wb.save(file)
        print(f'{file}已经处理完成')




